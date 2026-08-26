#!/usr/bin/env python3
"""
run_campagne.py
---------------
Orchestrateur principal — lit campagne_propulseur.xlsx, lance chaque run
OpenFOAM, écrit les résultats en temps réel dans la feuille "Resultats".

Usage:
    python3 scripts/run_campagne.py [--xlsx path/to/campagne.xlsx] [--ncores 8]

Colonnes lues dans la feuille "Runs" :
    A: Run#  B: Va[m/s]  C: RPM  H: Statut ("À faire" → sera traité)

Colonnes écrites dans la feuille "Resultats" :
    A-L selon en-têtes définis par create_campagne_template.py
"""

import argparse
import os
import sys
import subprocess
import shutil
import math
import datetime

import pandas as pd
import openpyxl

# ─────────────────────────────────────────────────────────────
# Chemins par défaut (relatifs à ce script)
# ─────────────────────────────────────────────────────────────
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
TP_ROOT    = os.path.normpath(os.path.join(SCRIPT_DIR, ".."))
XLSX_DEFAULT = os.path.join(TP_ROOT, "post", "campagne_propulseur.xlsx")
CASE_TEMPLATE = os.path.join(TP_ROOT, "cases", "propeller_mrf")
CASES_DIR     = os.path.join(TP_ROOT, "cases")
POST_DIR      = os.path.join(TP_ROOT, "post")
LOGS_DIR      = os.path.join(TP_ROOT, "logs")
FOAM_BASHRC   = "/usr/lib/openfoam/openfoam2412/etc/bashrc"

sys.path.insert(0, SCRIPT_DIR)
from patch_case import patch, foam_env

# ─────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────

def log(msg):
    ts = datetime.datetime.now().strftime("%H:%M:%S")
    print(f"[{ts}] {msg}", flush=True)


def run_shell(cmd, logfile=None, env=None):
    """Lance une commande bash, écrit dans logfile si fourni."""
    with open(logfile, "a") if logfile else open(os.devnull, "w") as lf:
        result = subprocess.run(
            ["bash", "-c", cmd],
            stdout=lf, stderr=subprocess.STDOUT,
            env=env or os.environ.copy(),
        )
    return result.returncode


def mean_last_frac(lst, frac=0.30):
    """Moyenne des derniers `frac` % de la liste."""
    if not lst:
        return float("nan")
    n = max(1, int(len(lst) * frac))
    return sum(lst[-n:]) / n


# ─────────────────────────────────────────────────────────────
# Lecture des résultats depuis log.pimpleFoam
# ─────────────────────────────────────────────────────────────

def parse_results(case_dir):
    log_path = os.path.join(case_dir, "log.pimpleFoam")
    if not os.path.isfile(log_path):
        return None

    J_arr, Kt_arr, Kq_arr, eta_arr = [], [], [], []
    res_p = []
    t = None

    with open(log_path) as f:
        for line in f:
            if line.startswith("Time = "):
                try: t = float(line.split()[2])
                except: pass
            if "Advance coefficient, J"    in line:
                try: J_arr.append(float(line.split(":")[1]))
                except: pass
            if "Thrust coefficient, Kt"    in line:
                try: Kt_arr.append(float(line.split(":")[1]))
                except: pass
            if "Torque coefficient, 10*Kq" in line:
                try: Kq_arr.append(float(line.split(":")[1]))
                except: pass
            if "Efficiency, etaO"          in line:
                try: eta_arr.append(float(line.split(":")[1]))
                except: pass
            if "Solving for p" in line and "GAMG" in line:
                try: res_p.append(float(line.split("Final residual = ")[1].split(",")[0]))
                except: pass

    if not Kt_arr:
        return None

    return {
        "J_moy":      round(mean_last_frac(J_arr),   4),
        "Kt_moy":     round(mean_last_frac(Kt_arr),  4),
        "Kq_moy":     round(mean_last_frac(Kq_arr),  4),
        "eta_moy":    round(mean_last_frac(eta_arr), 4),
        "Kt_final":   round(Kt_arr[-1],  4),
        "Kq_final":   round(Kq_arr[-1],  4),
        "eta_final":  round(eta_arr[-1], 4),
        "res_p":      f"{res_p[-1]:.3e}" if res_p else "N/A",
        "t_final":    t,
    }


# ─────────────────────────────────────────────────────────────
# Écriture dans la feuille Resultats (openpyxl — n'écrase pas les formules)
# ─────────────────────────────────────────────────────────────

def write_result(xlsx_path, run_num, Va, RPM, res):
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb["Resultats"]

    # Trouver la prochaine ligne vide (après en-tête)
    next_row = ws.max_row + 1
    if next_row <= 1:
        next_row = 2

    J = res["J_moy"]
    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")

    ws.cell(next_row,  1, value=run_num)
    ws.cell(next_row,  2, value=Va)
    ws.cell(next_row,  3, value=RPM)
    ws.cell(next_row,  4, value=J)
    ws.cell(next_row,  5, value=res["Kt_moy"])
    ws.cell(next_row,  6, value=res["Kq_moy"])
    ws.cell(next_row,  7, value=res["eta_moy"])
    ws.cell(next_row,  8, value=res["Kt_final"])
    ws.cell(next_row,  9, value=res["Kq_final"])
    ws.cell(next_row, 10, value=res["eta_final"])
    ws.cell(next_row, 11, value=res["res_p"])
    ws.cell(next_row, 12, value=now)

    wb.save(xlsx_path)
    log(f"Résultats écrits → Resultats ligne {next_row}")


def update_run_status(xlsx_path, run_num, status):
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb["Runs"]
    for row in ws.iter_rows(min_row=2):
        if row[0].value == run_num:
            row[7].value = status   # colonne H = Statut
            break
    wb.save(xlsx_path)


# ─────────────────────────────────────────────────────────────
# Run d'un point de fonctionnement
# ─────────────────────────────────────────────────────────────

def run_one(run_num, Va, RPM, ncores, end_time, xlsx_path):
    case_name = f"propeller_J{Va:.2f}Va_{int(RPM)}rpm".replace(".", "p")
    case_dir  = os.path.join(CASES_DIR, case_name)
    run_log   = os.path.join(LOGS_DIR, f"run_{case_name}.log")

    log(f"── Run #{run_num} : Va={Va} m/s  RPM={RPM}  → {case_name}")
    update_run_status(xlsx_path, run_num, "En cours…")

    # 1. Copier le template de cas
    if os.path.exists(case_dir):
        shutil.rmtree(case_dir)
    shutil.copytree(CASE_TEMPLATE, case_dir)

    # 2. Patcher Va et RPM
    patch(case_dir, Va, RPM)

    # Ajuster endTime
    env = foam_env()
    subprocess.run(
        ["bash", "-c",
         f". {FOAM_BASHRC} 2>/dev/null; "
         f"foamDictionary -entry endTime -set {end_time} {case_dir}/system/controlDict"],
        capture_output=True, env=env,
    )

    # 3. Lancer le calcul
    foam_cmd = (
        f". {FOAM_BASHRC} 2>/dev/null; "
        f". ${{WM_PROJECT_DIR}}/bin/tools/RunFunctions; "
        f"cd {case_dir}; "
        f"restore0Dir; "
        f"runApplication decomposePar; "
        f"runParallel -n {ncores} pimpleFoam; "
        f"runApplication reconstructPar"
    )
    rc = run_shell(foam_cmd, logfile=run_log, env=env)

    if rc != 0:
        log(f"[ERR] Run #{run_num} échoué (code {rc}) — voir {run_log}")
        update_run_status(xlsx_path, run_num, f"ERREUR (code {rc})")
        return

    # 4. Extraire et écrire les résultats
    res = parse_results(case_dir)
    if res:
        write_result(xlsx_path, run_num, Va, RPM, res)
        update_run_status(xlsx_path, run_num, "OK ✓")
        log(f"Run #{run_num} OK — J={res['J_moy']:.4f}  Kt={res['Kt_moy']:.4f}  "
            f"eta={res['eta_moy']*100:.1f}%")
    else:
        update_run_status(xlsx_path, run_num, "Résultats manquants")
        log(f"[WARN] Run #{run_num} : log pimpleFoam sans coefficients propulseur")


# ─────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Orchestrateur campagne propulseur")
    parser.add_argument("--xlsx",   default=XLSX_DEFAULT, help="Chemin du fichier campagne")
    parser.add_argument("--ncores", type=int, default=8,  help="Nombre de cœurs MPI")
    parser.add_argument("--only",   type=int, nargs="+",  help="Lancer seulement ces Run#")
    args = parser.parse_args()

    if not os.path.isfile(args.xlsx):
        print(f"[ERR] Fichier non trouvé : {args.xlsx}")
        print("      Lance d'abord : python3 scripts/create_campagne_template.py")
        sys.exit(1)

    # Lire la feuille Runs
    df = pd.read_excel(args.xlsx, sheet_name="Runs", header=0, engine="openpyxl")
    # Les colonnes avec formules sont lues comme NaN → recalcul côté Python
    df.columns = [c.strip().split("\n")[0] for c in df.columns]

    log(f"Campagne : {args.xlsx}")
    log(f"Runs définis : {len(df)}")

    for _, row in df.iterrows():
        try:
            run_num  = int(row.get("Run #", 0))
            Va       = float(row.get("Va [m/s]", 0))
            RPM      = float(row.get("RPM", 0))
            end_time = float(row.get("endTime [s]", 0.02))
            statut   = str(row.get("Statut", "")).strip()
        except (ValueError, TypeError):
            continue

        if args.only and run_num not in args.only:
            continue

        if statut not in ("À faire", "A faire", "à faire", "a faire", "TODO"):
            log(f"Run #{run_num} ignoré (Statut='{statut}')")
            continue

        run_one(run_num, Va, RPM, args.ncores, end_time, args.xlsx)

    log("═══ Campagne terminée ═══")
    log(f"Résultats → ouvrir {args.xlsx}")


if __name__ == "__main__":
    main()
