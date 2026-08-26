#!/usr/bin/env python3
"""
interpolate_courbe.py
---------------------
Lit la feuille "Resultats" du fichier campagne_propulseur.xlsx,
effectue une interpolation par spline cubique (scipy) sur Kt, 10Kq, etaO
en fonction de J, et écrit les points interpolés dans la feuille "Courbe".

Le graphe LibreOffice se met à jour automatiquement à la réouverture.

Usage:
    python3 scripts/interpolate_courbe.py [--xlsx path] [--npoints 100]
"""

import argparse
import os
import numpy as np
import openpyxl
from openpyxl.styles import PatternFill, Alignment, Border, Side
from scipy.interpolate import CubicSpline, interp1d

SCRIPT_DIR   = os.path.dirname(os.path.abspath(__file__))
TP_ROOT      = os.path.normpath(os.path.join(SCRIPT_DIR, ".."))
XLSX_DEFAULT = os.path.join(TP_ROOT, "post", "campagne_propulseur.xlsx")

THIN = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)

def style_data(cell, row_idx):
    cell.fill      = PatternFill("solid", fgColor="FEF9E7" if row_idx % 2 == 0 else "FFFFFF")
    cell.alignment = Alignment(horizontal="center")
    cell.border    = THIN
    cell.number_format = "0.0000"


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--xlsx",    default=XLSX_DEFAULT)
    parser.add_argument("--npoints", type=int, default=80,
                        help="Nombre de points interpolés sur la courbe eau libre")
    args = parser.parse_args()

    if not os.path.isfile(args.xlsx):
        print(f"[ERR] Fichier introuvable : {args.xlsx}")
        return

    # ── Lecture des résultats bruts ──────────────────────────
    import pandas as pd
    df = pd.read_excel(args.xlsx, sheet_name="Resultats", header=0, engine="openpyxl")
    df.columns = [str(c).strip() for c in df.columns]

    # Colonnes attendues : J [-], Kt_moy [-], 10Kq_moy [-], etaO_moy [-]
    col_J   = [c for c in df.columns if "J"   in c and "[-]" in c][0]
    col_Kt  = [c for c in df.columns if "Kt_moy" in c][0]
    col_Kq  = [c for c in df.columns if "10Kq_moy" in c or "Kq_moy" in c][0]
    col_eta = [c for c in df.columns if "etaO_moy" in c][0]

    df_clean = df[[col_J, col_Kt, col_Kq, col_eta]].dropna()
    df_clean = df_clean.sort_values(col_J).reset_index(drop=True)

    if len(df_clean) < 3:
        print(f"[WARN] Pas assez de points ({len(df_clean)}) pour interpoler. "
              "Il faut au minimum 3 runs converged.")
        # Écrire quand même les points bruts
        J_interp  = df_clean[col_J].values
        Kt_interp = df_clean[col_Kt].values
        Kq_interp = df_clean[col_Kq].values
        eta_interp = df_clean[col_eta].values
    else:
        J_raw   = df_clean[col_J].values
        Kt_raw  = df_clean[col_Kt].values
        Kq_raw  = df_clean[col_Kq].values
        eta_raw = df_clean[col_eta].values

        J_interp = np.linspace(J_raw.min(), J_raw.max(), args.npoints)

        # CubicSpline si assez de points, linéaire sinon
        if len(J_raw) >= 4:
            cs_Kt  = CubicSpline(J_raw, Kt_raw,  extrapolate=False)
            cs_Kq  = CubicSpline(J_raw, Kq_raw,  extrapolate=False)
            cs_eta = CubicSpline(J_raw, eta_raw,  extrapolate=False)
        else:
            cs_Kt  = interp1d(J_raw, Kt_raw,  kind="linear", fill_value="extrapolate")
            cs_Kq  = interp1d(J_raw, Kq_raw,  kind="linear", fill_value="extrapolate")
            cs_eta = interp1d(J_raw, eta_raw,  kind="linear", fill_value="extrapolate")

        Kt_interp  = cs_Kt(J_interp)
        Kq_interp  = cs_Kq(J_interp)
        eta_interp = cs_eta(J_interp)

        # Optionnel : ajouter la valeur J_max_eta (rendement max)
        i_max = int(np.argmax(eta_interp))
        print(f"[INFO] ηO max = {eta_interp[i_max]*100:.1f}%  à J = {J_interp[i_max]:.4f}")

    # ── Écriture dans la feuille Courbe ─────────────────────
    wb = openpyxl.load_workbook(args.xlsx)
    ws = wb["Courbe"]

    # Effacer les anciennes données (lignes 3 à 102)
    for row in range(3, 103):
        for col in range(1, 5):
            ws.cell(row=row, column=col, value=None)

    for i, (j, kt, kq, eta) in enumerate(
            zip(J_interp, Kt_interp, Kq_interp, eta_interp), start=3):
        ws.cell(i, 1, round(float(j),   5))
        ws.cell(i, 2, round(float(kt),  5))
        ws.cell(i, 3, round(float(kq),  5))
        ws.cell(i, 4, round(float(eta), 5))
        for col in range(1, 5):
            style_data(ws.cell(i, col), i)

    # Note de mise à jour
    now = __import__("datetime").datetime.now().strftime("%Y-%m-%d %H:%M")
    ws["A55"] = f"Mis à jour le {now} — {len(J_interp)} points, scipy CubicSpline"

    wb.save(args.xlsx)
    print(f"[OK] {len(J_interp)} points interpolés écrits → feuille 'Courbe'")
    print(f"[OK] Fichier sauvegardé : {args.xlsx}")
    print("     Rouvrir dans LibreOffice pour voir le graphe mis à jour.")


if __name__ == "__main__":
    main()
