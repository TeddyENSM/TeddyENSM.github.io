#!/usr/bin/env python3
"""
create_campagne_template.py
---------------------------
Génère le fichier campagne_propulseur.xlsx :

  Feuille "Runs"      : tu remplis Va [m/s] et RPM, le reste est auto
  Feuille "Resultats" : remplie par run_campagne.py (Kt, Kq, etaO, J)
  Feuille "Courbe"    : formules d'interpolation + graphe eau libre

Séparateur décimal : point (standard Python/OpenFOAM).
LibreOffice : Format > Cellules > Nombre si tu vois des virgules.
"""

import os
import openpyxl
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side)
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.utils import get_column_letter

OUT = os.path.join(os.path.dirname(__file__), "..", "post", "campagne_propulseur.xlsx")
OUT = os.path.normpath(OUT)

wb = openpyxl.Workbook()

# ─────────────────────────────────────────────────────────────
# Styles
# ─────────────────────────────────────────────────────────────
HDR_BLUE   = PatternFill("solid", fgColor="1A5276")
HDR_GREEN  = PatternFill("solid", fgColor="1E8449")
HDR_ORANGE = PatternFill("solid", fgColor="9C640C")
ALT_BLUE   = PatternFill("solid", fgColor="D6EAF8")
ALT_NONE   = PatternFill("solid", fgColor="FFFFFF")
WHITE_BOLD = Font(color="FFFFFF", bold=True)
THIN = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)

def style_header(cell, fill):
    cell.fill  = fill
    cell.font  = WHITE_BOLD
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = THIN

def style_data(cell, row_idx):
    cell.fill   = ALT_BLUE if row_idx % 2 == 0 else ALT_NONE
    cell.alignment = Alignment(horizontal="center")
    cell.border = THIN

# ─────────────────────────────────────────────────────────────
# Feuille 1 — Runs
# ─────────────────────────────────────────────────────────────
ws_runs = wb.active
ws_runs.title = "Runs"

headers_runs = [
    "Run #",
    "Va [m/s]",
    "RPM",
    "omega [rad/s]\n=RPM*2π/60",
    "J = Va/(n*D)\n(D=0.254 m)",
    "ncores",
    "endTime [s]",
    "Statut",
    "Notes",
]
ws_runs.row_dimensions[1].height = 40
for col, h in enumerate(headers_runs, 1):
    c = ws_runs.cell(row=1, column=col, value=h)
    style_header(c, HDR_BLUE)

# Paramètres fixes
D = 0.254   # diamètre hélice [m]
N_CORES = 8
END_TIME = 0.02

# Points de fonctionnement typiques (courbe eau libre)
#  Va de 0 à Va_max, RPM fixe = 25.15 tr/s = 1509 RPM
base_rpm = 1509
base_n   = base_rpm / 60   # tr/s

va_points = [0.5, 1.0, 1.5, 2.0, 2.5, 3.0, 4.0, 5.0]

for i, va in enumerate(va_points, start=1):
    row = i + 1
    omega = f"=C{row}*2*PI()/60"
    J     = f"=B{row}/(C{row}/60*{D})"
    ws_runs.cell(row=row, column=1, value=i)
    ws_runs.cell(row=row, column=2, value=va)
    ws_runs.cell(row=row, column=3, value=base_rpm)
    ws_runs.cell(row=row, column=4, value=omega)   # formule
    ws_runs.cell(row=row, column=5, value=J)        # formule
    ws_runs.cell(row=row, column=6, value=N_CORES)
    ws_runs.cell(row=row, column=7, value=END_TIME)
    ws_runs.cell(row=row, column=8, value="À faire")
    ws_runs.cell(row=row, column=9, value="")
    for col in range(1, 10):
        style_data(ws_runs.cell(row=row, column=col), i)

# Largeurs colonnes
for col, w in zip(range(1, 10), [7, 10, 10, 16, 16, 9, 13, 12, 25]):
    ws_runs.column_dimensions[get_column_letter(col)].width = w

# ─────────────────────────────────────────────────────────────
# Feuille 2 — Resultats  (remplie par run_campagne.py)
# ─────────────────────────────────────────────────────────────
ws_res = wb.create_sheet("Resultats")

headers_res = [
    "Run #", "Va [m/s]", "RPM", "J [-]",
    "Kt_moy [-]", "10Kq_moy [-]", "etaO_moy [-]",
    "Kt_final [-]", "10Kq_final [-]", "etaO_final [-]",
    "Residu_p_final", "Date run",
]
ws_res.row_dimensions[1].height = 35
for col, h in enumerate(headers_res, 1):
    c = ws_res.cell(row=1, column=col, value=h)
    style_header(c, HDR_GREEN)

for col, w in zip(range(1, 13), [7, 10, 8, 10, 12, 12, 12, 12, 12, 12, 14, 18]):
    ws_res.column_dimensions[get_column_letter(col)].width = w

# ─────────────────────────────────────────────────────────────
# Feuille 3 — Courbe  (interpolation + graphe)
# ─────────────────────────────────────────────────────────────
ws_courbe = wb.create_sheet("Courbe")

ws_courbe["A1"] = "Courbe eau libre — interpolée"
ws_courbe["A1"].font = Font(bold=True, size=13)

headers_courbe = ["J [-]", "Kt interp.", "10Kq interp.", "etaO interp."]
for col, h in enumerate(headers_courbe, 1):
    c = ws_courbe.cell(row=2, column=col, value=h)
    style_header(c, HDR_ORANGE)

# Lignes vides — remplies par interpolate_courbe.py
for row in range(3, 53):
    for col in range(1, 5):
        ws_courbe.cell(row=row, column=col, value="")

ws_courbe["A55"] = "⚙ Rempli automatiquement par : python3 scripts/interpolate_courbe.py"
ws_courbe["A55"].font = Font(italic=True, color="888888")

# Graphe scatter Kt/Kq/eta = f(J)
chart = ScatterChart()
chart.title       = "Courbe eau libre — pimpleFoam/AMI"
chart.style       = 10
chart.x_axis.title = "J [-]"
chart.y_axis.title = "Kt / 10Kq / ηO [-]"
chart.width  = 20
chart.height = 14

j_ref = Reference(ws_courbe, min_col=1, min_row=3, max_row=52)

for col_idx, (label, color) in enumerate([
    ("Kt",    "2471A3"),
    ("10Kq",  "E67E22"),
    ("etaO",  "27AE60"),
], start=2):
    y_ref = Reference(ws_courbe, min_col=col_idx, min_row=2, max_row=52)
    s = Series(y_ref, j_ref, title_from_data=True)
    s.graphicalProperties.line.solidFill = color
    s.graphicalProperties.line.width = 20000
    chart.series.append(s)

ws_courbe.add_chart(chart, "F2")

for col, w in zip(range(1, 5), [10, 14, 14, 14]):
    ws_courbe.column_dimensions[get_column_letter(col)].width = w

# ─────────────────────────────────────────────────────────────
# Sauvegarde
# ─────────────────────────────────────────────────────────────
wb.save(OUT)
print(f"[OK] Template créé : {OUT}")
print()
print("Workflow :")
print("  1. Ouvrir campagne_propulseur.xlsx dans LibreOffice Calc")
print("  2. Ajuster Va / RPM dans la feuille 'Runs'")
print("  3. Lancer : python3 scripts/run_campagne.py")
print("  4. Lancer : python3 scripts/interpolate_courbe.py")
print("  5. Rouvrir le fichier → feuille 'Courbe' mise à jour")
